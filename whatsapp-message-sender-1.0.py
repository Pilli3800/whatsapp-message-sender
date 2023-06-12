import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from tkinter import ttk
import pywhatkit
import pyautogui
from pynput.keyboard import Key, Controller
from win32 import win32clipboard
import pandas as pd

# Variables globales para almacenar las selecciones
excel_file = ""
selected_sheet = ""
selected_columns = []
selected_message_type = ""
message_text = ""
image_file = ""
selected_name_type = ""
file1_content = ""
file2_content = ""
type_name = ""

icon_path = ".\icon-message.ico"


def browse_excel():
    global excel_file
    filename = filedialog.askopenfilename(
        filetypes=[('Excel Files', '*.xlsx')])
    if filename:
        excel_file = filename.replace("/", "\\\\")
        excel_entry.delete(0, tk.END)
        excel_entry.insert(0, excel_file)
        load_excel_sheets()
        return excel_file
    else:
        messagebox.showwarning(
            "Advertencia", "No se seleccionó ningún archivo de Excel.")


def browse_image():
    global image_file
    filename = filedialog.askopenfilename(filetypes=[('JPEG Files', '*.jpeg')])
    if filename:
        image_file = filename.replace("/", "\\\\")
        image_entry.delete(0, tk.END)
        image_entry.insert(0, image_file)
        return image_file
    else:
        messagebox.showwarning(
            "Advertencia", "No se seleccionó ningún archivo de imagen JPEG.")


def browse_file(file_var):
    filename = filedialog.askopenfilename(filetypes=[('Text files', '*.txt')])
    if filename:
        file_var.set(filename)
        with open(filename, 'r', encoding='utf-8', errors='ignore') as file:
            if file_var == file1_var:
                global file1_content
                file1_content = file.read()
            elif file_var == file2_var:
                global file2_content
                file2_content = file.read()
    else:
        messagebox.showwarning(
            "Advertencia", "No se seleccionó ningún archivo.")


def load_excel_sheets():
    try:
        workbook = load_workbook(excel_file, read_only=True)
        sheet_names = workbook.sheetnames
        sheet_combobox['values'] = sheet_names
    except Exception as e:
        print(f"Error al cargar las hojas del archivo Excel: {e}")


def load_excel_columns():
    global selected_sheet

    selected_sheet = sheet_combobox.get()

    if not selected_sheet:
        return

    try:
        workbook = load_workbook(excel_file, read_only=True)
        sheet = workbook[selected_sheet]
        max_column = sheet.max_column
        columns = [sheet.cell(
            row=1, column=i).value for i in range(1, max_column+1)]
        column1_combobox['values'] = columns
        column2_combobox['values'] = columns
        column3_combobox['values'] = columns
    except Exception as e:
        print(f"Error al cargar las columnas del archivo Excel: {e}")


def submit():
    global selected_sheet, selected_columns, selected_message_type, message_text, image_file, selected_name_type

    if not excel_file:
        messagebox.showwarning(
            "Advertencia", "Debe seleccionar un archivo de Excel.")
        return

    selected_sheet = sheet_combobox.get()
    if not selected_sheet:
        messagebox.showwarning(
            "Advertencia", "Debe seleccionar una hoja del Excel.")
        return

    selected_columns = [
        column1_combobox.get(),
        column2_combobox.get(),
        column3_combobox.get()
    ]

    selected_message_type = message_type_combobox.get()
    message_text = message_entry.get("1.0", tk.END).strip()

    if not selected_message_type:
        messagebox.showwarning(
            "Advertencia", "Debe seleccionar un tipo de mensaje.")
        return

    selected_name_type = name_type_combobox.get()

    # Realizar operaciones con el archivo y las selecciones

    # Mostrar los valores seleccionados
    messagebox.showinfo("Valores seleccionados", f"Archivo de Excel: {excel_file}\n\n"
                        f"Hoja seleccionada: {selected_sheet}\n\n"
                        f"Columnas seleccionadas: {', '.join(selected_columns)}\n\n"
                        f"Tipo de mensaje seleccionado: {selected_message_type}\n\n"
                        f"Archivo de imagen: {image_file if image_file else 'No seleccionado'}\n\n"
                        f"Tipo de nombres seleccionado: {selected_name_type}\n\n"
                        f"Mensaje a enviar:\n {file1_content}\n"
                        f"{message_text}\n"
                        f"{file2_content}")


def send():

    print(messagebox.askokcancel(
        message="No debe usar la PC mientras el envio de mensajes esta en curso.", title="Importante"))

    global selected_sheet, selected_columns, selected_message_type, message_text, image_file, selected_name_type

    if not excel_file:
        messagebox.showwarning(
            "Advertencia", "Debe seleccionar un archivo de Excel.")
        return

    selected_sheet = sheet_combobox.get()
    if not selected_sheet:
        messagebox.showwarning(
            "Advertencia", "Debe seleccionar una hoja del Excel.")
        return

    selected_columns = [
        column1_combobox.get(),
        column2_combobox.get(),
        column3_combobox.get()
    ]

    selected_message_type = message_type_combobox.get()
    message_text = message_entry.get("1.0", tk.END).strip()
    message_text = f"{file1_content}\n\n{message_text}\n\n{file2_content}"
    message_text_withname = message_entry.get("1.0", tk.END).strip()

    if not selected_message_type:
        messagebox.showwarning(
            "Advertencia", "Debe seleccionar un tipo de mensaje.")
        return

    selected_name_type = name_type_combobox.get()

    type_name = name_type_combobox.get()

    # Evaluar el tipo de mensaje

    if (selected_message_type == "Enviar Solo Mensaje"):
        send_only_message(excel_file, sheet_combobox.get(),
                          message_text, column1_combobox.get())

    if (selected_message_type == "Enviar Solo Mensaje Con Imagen"):
        send_only_message_image(excel_file, image_file, sheet_combobox.get(),
                                message_text, column1_combobox.get())

    if (selected_message_type == "Enviar Mensajes con Encabezado"):
        send_message_withname(excel_file, type_name, sheet_combobox.get(),
                              message_text_withname, column1_combobox.get(), column2_combobox.get())

    if (selected_message_type == "Enviar Mensajes con Encabezado e Imagen"):
        send_message_withname_image(excel_file, image_file, type_name, sheet_combobox.get(),
                                    message_text_withname, column1_combobox.get(), column2_combobox.get())
    if (selected_message_type == "Firmas"):
        send_message_sign(excel_file, type_name, sheet_combobox.get(),
                          message_text_withname, column1_combobox.get(), column2_combobox.get(), column3_combobox.get())


# Mensajes generales sin imagen ni nombre

def send_only_message(excel_file: str, sheet_name: str, message_text: str, number_column: str):
    data = pd.read_excel(
        excel_file, sheet_name=sheet_name)
    for i in range(len(data)):
        number = str(
            data.loc[i, number_column])
        send_whatsapp_message(
            message_text=message_text, cellphone=number)
        print(str(i + 1) + ". Mensaje enviado a : " + number + " correctamente.")
    print(messagebox.askokcancel(
        message="Se finalizó el envio de mensajes.", title="Atención"))


def send_whatsapp_message(message_text: str, cellphone: str):
    try:
        pywhatkit.sendwhatmsg_instantly(
            phone_no="+51" + cellphone,
            message=message_text,
            tab_close=True
        )
        time.sleep(6)

        pyautogui.click()
        time.sleep(2)

        keyboard.press(Key.enter)
        keyboard.release(Key.enter)

    except Exception as e:
        print(str(e))

# Mensajes generales con imagen


def send_only_message(excel_file: str, sheet_name: str, message_text: str, number_column: str):
    data = pd.read_excel(
        excel_file, sheet_name=sheet_name)
    for i in range(len(data)):
        number = str(
            data.loc[i, number_column])
        send_whatsapp_message(
            message_text=message_text, cellphone=number)
        print(str(i + 1) + ". Mensaje enviado a : " + number + " correctamente.")
    print(messagebox.askokcancel(
        message="Se finalizó el envio de mensajes.", title="Atención"))


def send_only_message_image(excel_file: str, image_file: str, sheet_name: str, message_text: str, number_column: str):
    data = pd.read_excel(
        excel_file, sheet_name=sheet_name)
    for i in range(len(data)):
        number = str(
            data.loc[i, number_column])
        send_whatsapp_message_image(image_file=image_file,
                                    message_text=message_text, cellphone=number)
        print(str(i + 1) + ". Mensaje enviado a : " + number + " correctamente.")
    print(messagebox.askokcancel(
        message="Se finalizó el envio de mensajes.", title="Atención"))


def send_whatsapp_message_image(image_file: str, message_text: str, cellphone: str):
    try:
        pywhatkit.sendwhats_image(
            "+51" + cellphone, image_file, message_text, 10, True, 3)
        time.sleep(6)
        pyautogui.click()
        time.sleep(2)
        keyboard.press(Key.enter)
        keyboard.release(Key.enter)
    except Exception as e:
        print(str(e))

# Mensajes con Nombre


def send_message_withname(excel_file: str, type_name: str, sheet_name: str, message_text: str, number_column: str, name_column: str):
    data = pd.read_excel(
        excel_file, sheet_name=sheet_name)
    if (type_name.strip() == "Ninguno"):
        new_message_top = f"{file1_content}\n\n 📣 Estimado(a)"
    else:
        new_message_top = f"{file1_content}\n\n 📣 " + type_name
    for i in range(len(data)):
        number = str(
            data.loc[i, number_column])
        name = str(
            data.loc[i, name_column])
        send_whatsapp_message(
            message_text=new_message_top + " *" + name + "*\n\n" + message_text + "\n\n" + file2_content, cellphone=number)
        print(str(i + 1) + ". Mensaje enviado a : " + number + " correctamente.")
    print(messagebox.askokcancel(
        message="Se finalizó el envio de mensajes.", title="Atención"))

# Mensajes con Nombre e Imagen


def send_message_withname_image(excel_file: str, image_file: str, type_name: str, sheet_name: str, message_text: str, number_column: str, name_column: str):
    data = pd.read_excel(
        excel_file, sheet_name=sheet_name)
    if (type_name.strip() == "Ninguno"):
        new_message_top = f"{file1_content}\n\n 📣 Estimado(a)"
    else:
        new_message_top = f"{file1_content}\n\n 📣 " + type_name
    for i in range(len(data)):
        number = str(
            data.loc[i, number_column])
        name = str(
            data.loc[i, name_column])
        send_whatsapp_message_image(image_file=image_file,
                                    message_text=new_message_top + " *" + name + "*\n\n" + message_text + "\n\n" + file2_content, cellphone=number)
        print(str(i + 1) + ". Mensaje enviado a : " + number + " correctamente.")
    print(messagebox.askokcancel(
        message="Se finalizó el envio de mensajes.", title="Atención"))


def send_message_sign(excel_file: str, type_name: str, sheet_name: str, message_text: str, number_column: str, name_column: str, sign_column: str):
    data = pd.read_excel(
        excel_file, sheet_name=sheet_name)
    if (type_name.strip() == "Ninguno"):
        new_message_top = f"{file1_content}\n\n 📣 Estimado(a)"
    else:
        new_message_top = f"{file1_content}\n\n 📣 " + type_name
    for i in range(len(data)):
        number = str(
            data.loc[i, number_column]).replace(".0", "")

        name = str(
            data.loc[i, name_column])
        sign = str(
            data.loc[i, sign_column]).replace(".0", "")
        send_whatsapp_message(
            message_text=new_message_top + " *" + name + "*\n\n"+"Se le informa que tiene *" + sign + "* documento(s) sin firmar.", cellphone=number)
        print(str(i + 1) + ". Mensaje enviado a : " + number + " correctamente.")
    print(messagebox.askokcancel(
        message="Se finalizó el envio de mensajes.", title="Atención"))


# Crear la ventana principal
root = tk.Tk()
root.title("Whatsapp Message Sender")
root.geometry("800x600")

# Seleccionar archivo de Excel
excel_label = tk.Label(root, text="Seleccionar archivo de Excel:")
excel_label.grid(row=0, column=0, padx=10, pady=10, sticky="e")
excel_entry = tk.Entry(root, width=50)
excel_entry.grid(row=0, column=1, padx=10, pady=10)
excel_button = tk.Button(root, text="Seleccionar", command=browse_excel)
excel_button.grid(row=0, column=2, padx=10, pady=10)

# Seleccionar hoja del Excel
sheet_label = tk.Label(root, text="Nombre de Hoja de Datos:")
sheet_label.grid(row=1, column=0, padx=10, pady=10, sticky="e")
sheet_combobox = ttk.Combobox(root, width=50)
sheet_combobox.grid(row=1, column=1, padx=10, pady=10)
sheet_combobox.bind("<<ComboboxSelected>>", lambda event: load_excel_columns())

# Seleccionar columnas
column1_label = tk.Label(root, text="Columna de Numeros")
column1_label.grid(row=2, column=0, padx=10, pady=10, sticky="e")
column1_combobox = ttk.Combobox(root, width=50)
column1_combobox.grid(row=2, column=1, padx=10, pady=10)

column2_label = tk.Label(root, text="Columna de Nombres:")
column2_label.grid(row=3, column=0, padx=10, pady=10, sticky="e")
column2_combobox = ttk.Combobox(root, width=50)
column2_combobox.grid(row=3, column=1, padx=10, pady=10)

column3_label = tk.Label(root, text="Columna de Firmas")
column3_label.grid(row=4, column=0, padx=10, pady=10, sticky="e")
column3_combobox = ttk.Combobox(root, width=50)
column3_combobox.grid(row=4, column=1, padx=10, pady=10)

# Seleccionar tipo de nombres
name_type_label = tk.Label(root, text="Tipo de Nombre:")
name_type_label.grid(row=6, column=0, padx=10, pady=10, sticky="e")
name_type_combobox = ttk.Combobox(root, width=50)
name_type_combobox['values'] = ["Dr(a)", "Licenciado(a)", "Ninguno"]
name_type_combobox.grid(row=6, column=1, padx=10, pady=10)

# Seleccionar tipo de mensaje
message_type_label = tk.Label(root, text="Tipo de Mensaje:")
message_type_label.grid(row=7, column=0, padx=10, pady=10, sticky="e")
message_type_combobox = ttk.Combobox(root, width=50)
message_type_combobox['values'] = [
    "Enviar Solo Mensaje",
    "Enviar Solo Mensaje Con Imagen",
    "Enviar Mensajes con Encabezado",
    "Enviar Mensajes con Encabezado e Imagen",
    "Firmas"
]
message_type_combobox.grid(row=7, column=1, padx=10, pady=10)

# Ingresar texto del mensaje
message_label = tk.Label(root, text="Ingresar el texto del mensaje:")
message_label.grid(row=8, column=0, padx=10, pady=10, sticky="e")
message_entry = tk.Text(root, height=5, width=50)
message_entry.grid(row=8, column=1, padx=10, pady=10)

# Seleccionar archivo de imagen
image_label = tk.Label(root, text="Seleccionar archivo de imagen (JPEG):")
image_label.grid(row=9, column=0, padx=10, pady=10, sticky="e")
image_entry = tk.Entry(root, width=50)
image_entry.grid(row=9, column=1, padx=10, pady=10)
image_button = tk.Button(root, text="Seleccionar", command=browse_image)
image_button.grid(row=9, column=2, padx=10, pady=10)

# Seleccionar archivo 1
file1_label = tk.Label(root, text="Seleccionar Encabezado de Mensaje:")
file1_label.grid(row=10, column=0, padx=10, pady=10, sticky="e")
file1_var = tk.StringVar()
file1_entry = tk.Entry(root, textvariable=file1_var, width=30)
file1_entry.grid(row=10, column=1, padx=10, pady=10)
file1_button = tk.Button(root, text="Seleccionar",
                         command=lambda: browse_file(file1_var))
file1_button.grid(row=10, column=2, padx=10, pady=10)

# Seleccionar archivo 2
file2_label = tk.Label(root, text="Seleccionar Pie de Mensaje:")
file2_label.grid(row=11, column=0, padx=10, pady=10, sticky="e")
file2_var = tk.StringVar()
file2_entry = tk.Entry(root, textvariable=file2_var, width=30)
file2_entry.grid(row=11, column=1, padx=10, pady=10)
file2_button = tk.Button(root, text="Seleccionar",
                         command=lambda: browse_file(file2_var))
file2_button.grid(row=11, column=2, padx=10, pady=10)

# Botón de vista previa
submit_button = tk.Button(root, text="Ver Vista Previa", command=submit)
submit_button.grid(row=12, column=0, columnspan=3, padx=10, pady=10)

# Botón de envío
send_button = tk.Button(root, text="Enviar Mensajes", command=send)
send_button.grid(row=12, column=1, columnspan=3, padx=10, pady=10)

root.iconbitmap(icon_path)
root.resizable(False, False)
# Ejecutar el bucle principal de la aplicación
root.mainloop()
